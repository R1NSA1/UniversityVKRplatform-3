import io
import os

import asyncpg
import openpyxl
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font

load_dotenv()

app = FastAPI(title="Admin Service", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

APPROVED_QUERY = """
SELECT
    u_student.full_name AS student_name,
    u_student.email AS student_email,
    t.title AS topic_title,
    u_teacher.full_name AS supervisor_name,
    a.status::text AS status
FROM applications a
JOIN topics t ON t.id = a.topic_id
JOIN users u_student ON u_student.id::text = a.student_id
JOIN users u_teacher ON u_teacher.id::text = t.teacher_id
WHERE a.status::text = 'approved'
ORDER BY a.created_at DESC
"""


async def get_db_connection():
    return await asyncpg.connect(
        user=os.getenv("POSTGRES_USER", "postgres"),
        password=os.getenv("POSTGRES_PASSWORD"),
        database=os.getenv("POSTGRES_DB", "vkr_main"),
        host=os.getenv("POSTGRES_HOST", "postgres"),
        port=int(os.getenv("POSTGRES_PORT", "5432")),
    )


async def fetch_approved_applications():
    conn = await get_db_connection()
    try:
        rows = await conn.fetch(APPROVED_QUERY)
        return [dict(row) for row in rows]
    except asyncpg.UndefinedTableError:
        return []
    finally:
        await conn.close()


def build_workbook(applications: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Студент", "Email", "Тема ВКР", "Руководитель", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, app in enumerate(applications, 2):
        ws.cell(row=row_idx, column=1, value=app.get("student_name", ""))
        ws.cell(row=row_idx, column=2, value=app.get("student_email", ""))
        ws.cell(row=row_idx, column=3, value=app.get("topic_title", ""))
        ws.cell(row=row_idx, column=4, value=app.get("supervisor_name", ""))
        status = app.get("status", "")
        ws.cell(
            row=row_idx,
            column=5,
            value="Утверждена" if status == "approved" else status,
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/health")
async def health():
    return {"status": "ok", "service": "admin"}


@app.get("/api/admin/applications")
async def get_applications():
    try:
        applications = await fetch_approved_applications()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось загрузить заявки: {exc}") from exc
    return {"applications": applications}


@app.get("/api/admin/export/excel")
async def export_excel():
    try:
        applications = await fetch_approved_applications()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось выгрузить Excel: {exc}") from exc

    output = build_workbook(applications)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=applications.xlsx"},
    )
